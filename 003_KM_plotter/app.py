'''
App made by:
    Eduardo Reyes Alvarez, Ph.D.
Contact:
    eduardo_reyes09@hotmail.com

App version: 
    V04 (Dec 15, 2023): Fourth partial version of the app. A new approach to create the subgrouping
                        widgets, their containers and response elements has been found and partially
                        implemented. Reviewing the logging is still pending.
'''
###################################################################################################

# Import required libraries

import io
import logging
from typing import List
from collections import OrderedDict

import numpy as np
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
import altair as alt 

import streamlit as st
from streamlit_searchbox import st_searchbox

from lifelines import KaplanMeierFitter
from lifelines.plotting import add_at_risk_counts

###################################################################################################

# App configuration and layout
st.set_page_config(
    page_title="Tool 003 - App by Eduardo",
    page_icon=":chart_with_downwards_trend:",
    layout="wide",
    initial_sidebar_state="expanded")

# Title
st.title("Interactive Kaplan-Meier plot generator")
st.markdown('<hr style="margin-top: +2px; margin-bottom: +2px;">', unsafe_allow_html=True)

# Sidebar - Initial widgets
with st.sidebar:
    uploaded_files = st.file_uploader(label="Upload a clinical file (and optionally, a RNA file)", 
                                      type=["txt"], accept_multiple_files=True)
    start_button = st.button(label="Begin", type="secondary")
    restart_button = st.button(label="Start over", type="secondary")

###################################################################################################

# Function to setup the logging configuration
def logging_setup():
     
     # Check if logging has already been initialized
     if "log_created" not in st.session_state:
        
        # Configure the logging settings
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

        # Create a logger
        logger = logging.getLogger()

        # Clear the existing log file or create a new empty one
        open('MyLog.txt', 'w').close()

        # Create a file handler
        file_handler = logging.FileHandler('MyLog.txt')
        file_handler.setLevel(logging.INFO)

        # Create a formatter and add it to the file handler
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)

        # Add the file handler to the logger
        logger.addHandler(file_handler)

        # Log an initial message
        logger.info(f"Log file created or cleared. \n")

        # Update the run variable so this set up process gets executed only once
        st.session_state["logger"] = logger
        st.session_state["log_created"] = True

###################################################################################################

# Function to prompt the user to upload the files if not already uploaded
def load_input_files(uploaded_files):

    # Check if one or two files were uploaded
    clinical_file = next((file for file in uploaded_files if file.name == 'clinical.txt'), None)
    RNA_file = next((file for file in uploaded_files if file.name == 'RNA.txt'), None)
       
    # Load the files available
    if clinical_file is not None:
        
        # Initialize the logging file 
        logging_setup()
        logger = st.session_state["logger"]

        # Load, log and save to the session state the mandatory file
        df_clinical = pd.read_csv(clinical_file, sep="\t", comment="#")
        logger.info(f"File found: clinical.txt \n")
        st.session_state["df_clinical"] = df_clinical

        # Load, log and save to the session state the optional file if it exists
        if RNA_file is not None:
            df_RNA = pd.read_csv(RNA_file, sep="\t")
            logger.info(f"File found: RNA.txt \n")
            st.session_state["df_RNA"] = df_RNA            
    else:
        st.sidebar.error("A clinical file is required!")
        st.stop()

###################################################################################################

# This function searches OS/PFS/RFS/DFS _STATUS and _MONTHS columns in the clinical data
# This function transposes the RNA dataset to have gene names as columns and patient IDs as rows (like the clinical data)

def file_preprocessing():

    # Get the df(s) and the logger
    df_clinical = st.session_state.get("df_clinical")
    df_RNA = st.session_state.get("df_RNA", None)
    logger = st.session_state.get("logger")

    ################### Processing for the clinical dataframe ##################

    # Log the original dataframe
    logger.info(f"Preview of the original clinical dataset: \n {df_clinical.iloc[:15, :10].to_string()} \n")
    logger.info(f"Data types of columns in the original clinical dataset: \n {df_clinical.dtypes.to_string()} \n\n")

    # Prepare the variable with the reordered column names
    clinical_columns_main = ["PATIENT_ID"]
    time_to_event_options = []
    event_observation_options = []

    # Search for possible metric measurements (Overall/Recurrence-Free/Progression-Free/Disease-Free/Disease-specific Survival
    for metric in ["OS", "RFS", "PFS", "DFS", "DSS"]:
        if metric+"_MONTHS" in df_clinical.columns and metric+"_STATUS" in df_clinical.columns:
            clinical_columns_main.append(metric+"_MONTHS")
            time_to_event_options.append(metric+"_MONTHS")
            clinical_columns_main.append(metric+"_STATUS")
            event_observation_options.append(metric+"_STATUS")

    # Search for a column of Vital status or Cause of Death (this is optional and may provide useful information)
    for extra_metric in ["VITAL_STATUS", "CAUSE_OF_DEATH"]:
        if extra_metric in df_clinical.columns:
            clinical_columns_main.append(extra_metric)
            event_observation_options.append(extra_metric)

    # Order alphabetically the remaining columns
    clinical_columns_extra = [col for col in df_clinical.columns if col not in clinical_columns_main]
    clinical_columns_extra.sort()

    # Apply the re-ordering to the df
    clinical_columns_ordered = clinical_columns_main + clinical_columns_extra
    df_clinical = df_clinical[clinical_columns_ordered] 

    # Log the re-arranged dataframe
    logger.info(f"Preview of the pre-processed clinical dataset: \n {df_clinical.iloc[:15, :10].to_string()} \n")
    logger.info(f"Data types of columns in the pre-processed clinical dataset: \n {df_clinical.dtypes.to_string()} \n\n")

    ################### Processing for the RNA dataframe ###################

    # If an RNA file was uploaded, then the df is not empty
    if df_RNA is not None:
        # Log the original dataframe
        logger.info(f"Preview of the original RNA dataset: \n {df_RNA.iloc[:15, :10].to_string()} \n")
        logger.info(f"Data types of some columns in the original RNA dataset: \n {df_RNA.iloc[:, :10].dtypes.to_string()} \n\n")
        
        # Drop the "Entrez_Gene_Id" column if exists
        if "Entrez_Gene_Id" in df_RNA.columns:
            df_RNA.drop("Entrez_Gene_Id", axis=1, inplace=True)
        
        # Rename the "Hugo_Symbol" column to "PATIENT_ID" as it appears in the clinical df
        df_RNA.rename(columns={"Hugo_Symbol": "PATIENT_ID"}, inplace=True)
        
        # Transpose the dataframe, making the content of the "PATIENT_ID" column the new column names
        df_RNA = df_RNA.set_index("PATIENT_ID").T
        
        # Sort the gene names alphabetically
        df_RNA.sort_index(axis=1, inplace=True)
        
        # Reset the index to a numerical index
        df_RNA = df_RNA.reset_index().rename_axis("", axis="columns")
        
        # Rename the "index" column to "PATIENT_ID"
        df_RNA.rename(columns={"": "PATIENT_ID"}, inplace=True)
        
        # Sort Patient IDs and reset the index
        df_RNA = df_RNA.sort_values("PATIENT_ID").reset_index(drop=True)

        # Log the re-arranged dataframe
        logger.info(f"Preview of the pre-processed RNA dataset: \n {df_RNA.iloc[:15, :10].to_string()} \n")
        logger.info(f"Data types of some columns in the pre-processed RNA dataset: \n {df_RNA.iloc[:, :10].dtypes.to_string()} \n\n")

    ###################

    # Save the neccesary data for the next steps in the session state
    st.session_state.update({"df_clinical": df_clinical,
                             "df_RNA": df_RNA,
                             "time_to_event_options": time_to_event_options,
                             "event_observation_options": event_observation_options})

    # Also add the options of genes if a RNA file was uploaded
    if df_RNA is not None:
        st.session_state["gene_list"] = tuple(df_RNA.columns[1:].tolist())

###################################################################################################

# Main function to prepare and display the interactive widgets and subwidgets
def widget_preparation():

    # Get required variables from the session state
    time_to_event_options = st.session_state.get("time_to_event_options")
    event_observation_options = st.session_state.get("event_observation_options")
    logger = st.session_state.get("logger")

    logger.info(f"---------------User interaction with the widgets starts here--------------- \n")

    # Create the layout for the widgets
    col_1_row_1, col_2_row_1 = st.columns(2, gap="medium")
    col_1_row_2, col_2_row_2 = st.columns(2, gap="medium")
    st.markdown('<hr style="margin-top: -15px; margin-bottom: -15px;">', unsafe_allow_html=True)
    col_1_row_3, col_2_row_3, col_3_row_3, col_4_row_3 = st.columns([2, 2, 1.35, 1.15], gap="medium")
    st.markdown('<hr style="margin-top: +10px; margin-bottom: +10px;">', unsafe_allow_html=True)
    col_1_row_14, col_2_row_14, col_3_row_14 = st.columns([0.5, 9, 0.5], gap="medium")
    st.markdown('<hr style="margin-top: +2px; margin-bottom: +2px;">', unsafe_allow_html=True)
    col_1_row_15, col_2_row_15, col_3_row_15, col_4_row_15 = st.columns(4, gap="small")
    
    # Save the columns and containers in the session state
    widget_and_output_areas = [col_1_row_1, col_2_row_1,
                               col_1_row_2, col_2_row_2,
                               col_1_row_3, col_2_row_3, col_3_row_3, col_4_row_3,
                               col_1_row_14, col_2_row_14, col_3_row_14,
                               col_1_row_15, col_2_row_15, col_3_row_15, col_4_row_15]
    st.session_state["widget_and_output_areas"] = widget_and_output_areas
    
    # Time to event widget and callback function
    with col_1_row_1:
        time_to_event_dropdown = st.selectbox(label="Select the time-to-event column", 
                                options=["Click here to select..."] + time_to_event_options)
    if time_to_event_dropdown:
        time_to_event_dropdown_handler(time_to_event_dropdown)

    # Event observation widget and callback function
    with col_2_row_1:
        event_observation_dropdown = st.selectbox(label="Select the event observation column", 
                            options=["Click here to select..."] + event_observation_options)
    if event_observation_dropdown:
        event_observation_dropdown_handler(event_observation_dropdown)

    # Show widgets to generate+save the plot (with their callback functions), and to customize it
    with st.sidebar:
        generate_plot_button = st.button(label="Generate plot", type="primary")
        st.markdown('<hr style="margin-top: 1px; margin-bottom: 1px;">', unsafe_allow_html=True)
        
        # Customize the plot
        st.write("#### Customize your plot here  ⬇️⬇️")
        CI_checkbox = st.checkbox(label="Show Confidence Intervals", value=True)
        move_labels_checkbox = st.checkbox(label="Move legend to the side", value=False)
        at_risk_checkbox = st.checkbox(label="Show at-risk table", value=False)
        sample_percent_slider = st.slider(label="Datapoints to plot (%)", min_value=50, max_value=100, value=95)
    
    # Add the plot customization options to the session state and save button to enable it later
    st.session_state["CI_checkbox"] = CI_checkbox
    st.session_state["move_labels_checkbox"] = move_labels_checkbox
    st.session_state["at_risk_checkbox"] = at_risk_checkbox
    st.session_state["sample_fraction"] = sample_percent_slider / 100
    
    # Control the rerun of the plot+excel generation process as not all interactions should trigger it
    if generate_plot_button or "logged_figure" in st.session_state:
        KM_plot_area = col_2_row_14
        
        # When reruns occur due to other widgets, show the same plot unless the user wants a new one
        previous_plot = st.session_state.get("logged_figure", False)
        if previous_plot and not generate_plot_button:
            with KM_plot_area:
                
                st.image(previous_plot)
        else:
            KM_figure = pass_KM_parameters()
            with KM_plot_area:
                KM_plot_area.empty()
                st.pyplot(KM_figure)

        # In any case the function to save the files should be executed (it handles logged files)
        save_KM_results(generate_plot_button)

    #################################

    # # Create 5 subwidget areas (more subwidgets) to use a maximum of 5 variables with this button (min 1)
    # subgroup_widget_areas = [{ 'dataset_dropdown' : widgets.Dropdown(options=['Click here to select...', 'clinical'] + (['RNA'] if df_RNA is not None else []),
    #                                                                  value='Click here to select...', description='Dataset:'),
    #                           'variables_dropdown' : widgets.Dropdown(options=['Click here to select...'] + list(df_clinical.columns[1:]), 
    #                                                                   value='Click here to select...', description='Variables:'),
    #                           'variables_combobox' : widgets.Combobox(options=list(df_RNA.columns[1:]), placeholder='Type gene of interest here', 
    #                                                                   description='Genes:') if df_RNA is not None else None,
    #                           'subgroup_number_slider' : widgets.IntSlider(min=1, max=10, description='Groups:', value=1) 
    #                           } for i in range(5)]

    # # Variables that need to be initialized with specific values/number of items
    # column_data = ["0", "1", "2", "3", "4"]
    # subgroup_boxes = ["0", "1", "2", "3", "4"]
    # KM_data_all = pd.DataFrame(columns=["0", "1", "2"])

###################################################################################################

# Function to display the output of time_to_event_dropdown (histogram)
def time_to_event_dropdown_handler(change):
    
    # Save the current selection to the session state
    st.session_state["time_to_event_selection"] = change

    # Get the required variables from the session state
    df_clinical = st.session_state.get("df_clinical")
    logger = st.session_state.get("logger")
    time_to_event_output = st.session_state["widget_and_output_areas"][2]

    # Clear the output area and return early if the default option is selected back 
    if change == "Click here to select...":
        with time_to_event_output: 
            time_to_event_output.empty()
        return

    # If the selection is a column on the clinical dataframe, display a histogram
    column_name = change
    if column_name in df_clinical.columns:
        time_column = df_clinical[column_name].dropna()
        logger.info(f"The user selected: {column_name}     Widget: time_to_event_dropdown. \n")
        logger.info(f"Original dtype of {column_name}: {df_clinical[column_name].dtype}     Dtype once removing NANs: {time_column.dtype} \n")

        # Make histogram of values with altair and handle exceptions
        if time_column.dtype != "object":
            altair_data = pd.DataFrame({column_name: time_column})

            chart1 = alt.Chart(altair_data).mark_bar(color='#BA4A00').encode(
                alt.X(column_name, type='quantitative', bin=alt.Bin(step=12)),
                alt.Y('count()', title='Patients'),
                ).properties(width=425, height=325
                ).configure_axis(labelColor="#2471A3")
            logger.info(f"A histogram was successfully made and displayed for: {column_name} \n")
        else:
            # If the column is not numeric, display a warning and stop the app
            info_str = "Warning: Column type is not numeric."
            logger.warning("User attention required: The time to event column may not be numerical. \n")
            
            with time_to_event_output:
                time_to_event_output.empty()
                st.warning(info_str)
            st.stop()
    else:
        # If the column is not in the df, display an error and stop the app
        info_str = "Warning: Column not found in the dataframe."
        logger.error("User attention required: The time to event column name was not found in the df. \n")
        
        with time_to_event_output:
            time_to_event_output.empty()
            st.error(info_str)
        st.stop()

    # Clear the output area and display the histogram
    with time_to_event_output:
        time_to_event_output.empty()
        st.altair_chart(chart1)

###################################################################################################

# Function to display the output of event_observation_dropdown (bar chart)
def event_observation_dropdown_handler(change):

    # Save the current column selection to the session state
    st.session_state["event_observation_selection"] = change

    # Get the required variables from the session state
    df_clinical = st.session_state.get("df_clinical")
    logger = st.session_state.get("logger")
    event_observation_output_1 = st.session_state["widget_and_output_areas"][3]
    event_observation_output_2 = st.session_state["widget_and_output_areas"][4]
    event_observation_output_3 = st.session_state["widget_and_output_areas"][5]
    event_observation_output_4 = st.session_state["widget_and_output_areas"][6]

    # Clear the output and return early if the default option is selected back 
    if change == "Click here to select...":
        event_observation_output_1.empty()
        event_observation_output_2.empty()
        event_observation_output_3.empty()
        event_observation_output_4.empty()
        return
        
    # If the selection is a column on the clinical dataframe, display a bar chart
    column_name = change
    if column_name in df_clinical.columns:
        event_column = df_clinical[column_name]
        logger.info(f"The user selected: {column_name}     Widget: event_observation_dropdown. \n")
        logger.info(f"Dtype of {column_name}: {df_clinical[column_name].dtype}     Unique value counts: \n\t {event_column.value_counts(dropna=False).to_string()} \n")

        # Make a bar chart for unique values in the column and handle exceptions
        if event_column.dtype == "object":
            value_counts = event_column.value_counts(dropna=True)
            
            if df_clinical[column_name].nunique() > 10:
                logger.warning("User attention required: There may be something wrong with the event observation column as there are more than 15 unique values. \n")
            
            # Use altair to make the graph, using these pre-selected colors
            alt_colors = ['#76448A', '#B03A2E', '#1E8449', '#1F618D', '#922B21',  
                          '#ffff33', '#a65628', '#f781bf', '#999999', '#66c2a5']
            
            chart2 = alt.Chart(value_counts.reset_index()).mark_bar().encode(
                    alt.X(column_name, type='nominal', axis=alt.Axis(labelAngle=0)),
                    alt.Y('count', title="Patients"),
                    alt.Color(column_name, scale=alt.Scale(domain=list(value_counts.index), range=alt_colors))
                    ).properties(width=425, height=325
                    ).configure_axis(labelColor="#2471A3",
                    ).configure_legend(disable=True)
            logger.info(f"A bar chart was successfully made and displayed for: {column_name} \n")
        else:
            # If the column is not categorical, show a warning and stop the app
            info_str = "Warning: Column type is not categorical."
            logger.warning("User attention required: The event observation column may not be text-based. \n")
            
            with event_observation_output_1:
                event_observation_output_1.empty()
                st.warning(info_str)
            st.stop()
    else:
        # If the selection is not a column in the df, show an error and stop the app
        info_str = "Warning: Column not found in the dataframe."
        logger.error("User attention required: The event observation column name was not found in the df. \n")

        with event_observation_output_1:
            event_observation_output_1.empty()
            st.error(info_str)
        st.stop()

    # Clear the output area and display the bar chart
    with event_observation_output_1:
        event_observation_output_1.empty()
        st.altair_chart(chart2)
    
    # Get the unique values of the column currently selected    
    event_options = np.ndarray.tolist(df_clinical[change].unique())

    # Make subwidgets to specify the event to be observed so we can encode it in binary
    with event_observation_output_2:
        event_observation_output_2.empty()
        event_options_0 = st.multiselect(label="No event (0):", options=event_options)
    with event_observation_output_3:
        event_observation_output_3.empty()
        event_options_1 = st.multiselect(label="Event (1):", options=event_options)
    
    # Make a widget to ask the user if they want a curve for the whole dataset or divide it
    with event_observation_output_4:
        subgroup_buttons = st.radio(label="Make subgroups?", options=["None", "Using variable(s)"], index=0)
    if subgroup_buttons:
        if "logged_figure" in st.session_state:
            del st.session_state["logged_figure"]
        subgroup_buttons_handler(subgroup_buttons)
    
    # Save the selected events in the session state
    st.session_state["event_0"] = event_options_0
    st.session_state["event_1"] = event_options_1

###################################################################################################

# Function to display the output of subgroup_buttons (slider)
def subgroup_buttons_handler(change):
    
    # Save the subgroup selection to the session state 
    st.session_state["subgroup_buttons_selection"] = change

    # Get the required variables from the session state
    logger = st.session_state.get("logger")
    subgroup_buttons_output = st.session_state["widget_and_output_areas"][7]
    
    # If the user wants to make subgroups, ask the number of variables to use
    if change == 'Using variable(s)':
        logger.info(f"The user selected: Use variable(s)     Widget: subgroup_buttons \n")
        
        # Show a slider for 1 to 5 variables (to prevent crazy number of curves)
        with subgroup_buttons_output:
            variable_number_slider = st.slider(label="Number of variables:", 
                                               min_value=1, max_value=5, step=1, value=1)
        variable_number_slider_handler(variable_number_slider)   
    else:
        # If the user doesn't want to make subgroups, don't show the slider 
        logger.info(f"The user selected: No subgroups     Widget: subgroup_buttons \n")
        subgroup_buttons_output.empty()

###################################################################################################

# Function to display the output of variable_number_slider (widget+output areas)
def variable_number_slider_handler(change):
    
    # Get required variables from the session state
    df_clinical = st.session_state.get("df_clinical")
    df_RNA = st.session_state.get("df_RNA")
    logger = st.session_state.get("logger")
    
    # Repeat 1
    if change >= 1:
        # Create a list to store the column containers
        col_1_row_4, col_2_row_4, col_3_row_4 = st.columns([1,1,1])
        col_1_row_5, col_2_row_5 = st.columns([3,1])

        # Create and display the appropriate widgets
        with col_1_row_4:
            dataset_dropdown_1 = st.selectbox(label="Select a dataset:", 
                        options=['Click here to select...', 'clinical'] + (['RNA'] if df_RNA is not None else []),
                        index=0, key="dataset_dropdown_1")
            
        if dataset_dropdown_1 == 'clinical':
            with col_2_row_4:
                variable_dropdown_1 = st.selectbox(label="Select a variable:",
                                    options=['Click here to select...'] + list(df_clinical.columns[1:]),
                                    index=0, key="variable_dropdown_1")
            with col_3_row_4:
                subgroup_slider_1 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=10, value=1, step=1, key="subgroup_slider_1") 
        elif dataset_dropdown_1 == 'RNA':
            with col_2_row_4:
                variable_dropdown_1 = st_searchbox(search_function=search_genes, default=None, 
                     label="Type a gene name here", clear_on_submit=False, key="variable_dropdown_1")
            with col_3_row_4:
                subgroup_slider_1 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=10, value=1, step=1, key="subgroup_slider_1") 

    # Repeat 2
    if change >= 2:
        col_1_row_6, col_2_row_6, col_3_row_6 = st.columns([1,1,1])
        col_1_row_7, col_2_row_7 = st.columns([3,1])

                # Create and display the appropriate widgets
        with col_1_row_6:
            dataset_dropdown_2 = st.selectbox(label="Select a dataset:", 
                        options=['Click here to select...', 'clinical'] + (['RNA'] if df_RNA is not None else []),
                        index=0, key="dataset_dropdown_2")
            
        if dataset_dropdown_2 == 'clinical':
            with col_2_row_6:
                variable_dropdown_2 = st.selectbox(label="Select a variable:",
                                    options=['Click here to select...'] + list(df_clinical.columns[1:]),
                                    index=0, key="variable_dropdown_2")
            with col_3_row_6:
                subgroup_slider_2 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=10, value=1, step=1, key="subgroup_slider_2") 
        elif dataset_dropdown_2 == 'RNA':
            with col_2_row_6:
                variable_dropdown_2 = st_searchbox(search_function=search_genes, default=None, 
                     label="Type a gene name here", clear_on_submit=False, key="variable_dropdown_2")
            with col_3_row_6:
                subgroup_slider_2 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=10, value=1, step=1, key="subgroup_slider_2") 

    # Repeat 3       
    if change >= 3:
        col_1_row_8, col_2_row_8, col_3_row_8 = st.columns([1,1,1])
        col_1_row_9, col_2_row_9 = st.columns([3,1])

        # Create and display the appropriate widgets
        with col_1_row_8:
            dataset_dropdown_3 = st.selectbox(label="Select a dataset:", 
                        options=['Click here to select...', 'clinical'] + (['RNA'] if df_RNA is not None else []),
                        index=0, key="dataset_dropdown_3")
            
        if dataset_dropdown_3 == 'clinical':
            with col_2_row_8:
                variable_dropdown_3 = st.selectbox(label="Select a variable:",
                                    options=['Click here to select...'] + list(df_clinical.columns[1:]),
                                    index=0, key="variable_dropdown_3")
            with col_3_row_8:
                subgroup_slider_3 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=10, value=1, step=1, key="subgroup_slider_3") 
        elif dataset_dropdown_3 == 'RNA':
            with col_2_row_8:
                variable_dropdown_3 = st_searchbox(search_function=search_genes, default=None, 
                     label="Type a gene name here", clear_on_submit=False, key="variable_dropdown_3")
            with col_3_row_8:
                subgroup_slider_3 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=10, value=1, step=1, key="subgroup_slider_3") 
    # Repeat 4
    if change >= 4:
        col_1_row_10, col_2_row_10, col_3_row_10 = st.columns([1,1,1])
        col_1_row_11, col_2_row_11 = st.columns([3,1])

        # Create and display the appropriate widgets
        with col_1_row_10:
            dataset_dropdown_4 = st.selectbox(label="Select a dataset:", 
                        options=['Click here to select...', 'clinical'] + (['RNA'] if df_RNA is not None else []),
                        index=0, key="dataset_dropdown_4")
            
        if dataset_dropdown_4 == 'clinical':
            with col_2_row_10:
                variable_dropdown_4 = st.selectbox(label="Select a variable:",
                                    options=['Click here to select...'] + list(df_clinical.columns[1:]),
                                    index=0, key="variable_dropdown_4")
            with col_3_row_10:
                subgroup_slider_4 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=10, value=1, step=1, key="subgroup_slider_4") 
        elif dataset_dropdown_1 == 'RNA':
            with col_2_row_10:
                variable_dropdown_4 = st_searchbox(search_function=search_genes, default=None, 
                     label="Type a gene name here", clear_on_submit=False, key="variable_dropdown_4")
            with col_3_row_10:
                subgroup_slider_4 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=10, value=1, step=1, key="subgroup_slider_4") 
    # Repeat 5
    if change >= 5:
        col_1_row_12, col_2_row_12, col_3_row_12 = st.columns([1,1,1])
        col_1_row_13, col_2_row_13 = st.columns([3,1])

        # Create and display the appropriate widgets
        with col_1_row_12:
            dataset_dropdown_5 = st.selectbox(label="Select a dataset:", 
                        options=['Click here to select...', 'clinical'] + (['RNA'] if df_RNA is not None else []),
                        index=0, key="dataset_dropdown_5")
            
        if dataset_dropdown_5 == 'clinical':
            with col_2_row_12:
                variable_dropdown_5 = st.selectbox(label="Select a variable:",
                                    options=['Click here to select...'] + list(df_clinical.columns[1:]),
                                    index=0, key="variable_dropdown_5")
            with col_3_row_12:
                subgroup_slider_5 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=10, value=1, step=1, key="subgroup_slider_5") 
        elif dataset_dropdown_5 == 'RNA':
            with col_2_row_12:
                variable_dropdown_5 = st_searchbox(search_function=search_genes, default=None, 
                     label="Type a gene name here", clear_on_submit=False, key="variable_dropdown_5")
            with col_3_row_12:
                subgroup_slider_5 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=10, value=1, step=1, key="subgroup_slider_5") 

###################################################################################################

# Function to search a gene of interest in the RNA dataset (if provided) to make subgroups
def search_genes(searchterm: str) -> List[tuple[str, str]]:
    
    # Assuming st.session_state["gene_list"] is a list of gene names
    suggestions = [gene for gene in st.session_state["gene_list"] if searchterm.lower() in gene.lower()]
    
    # Returning a list of tuples where each tuple contains a label and a value
    return [(gene, gene) for gene in suggestions]

###################################################################################################

# Function to display the output of variables_dropdown and variables_combobox (plots)
# Reminder that this function has to work for both df_clinical and df_RNA
def variables_selection_handler(change, repeat):
    st.write("Feature 2 pending")
    return
    # Reset the subgroup_number_slider to remove any previous widget boxes
    subgroup_widget_areas[repeat]['subgroup_number_slider'].value = 1
    
    # Display empty space if the default value on the dropdown or combobox is selected
    if change == 'Click here to select...' or change == '':
        logger.info(f"The previous variable to make subgroups was de-selected. \n")
        subgroup_output_areas[repeat]['subgroup_maker1_info'].clear_output()
        subgroup_output_areas[repeat]['subgroup_maker2_info'].clear_output()
        return
    # If the user has selected a variable, do 4 steps: Apply event labels, find the column, plot, capture all columns of interest
    else:
        ##### Preparation part
        # Make these variables global so they can be accessed from other subwidgets
        global column_data, KM_data_1var, KM_data_all
        KM_data_1var = df_clinical.copy()
        
        # This is to avoid plotting while incomplete strings are being searched in the combobox
        if subgroup_widget_areas[repeat]['variables_combobox'] is not None and subgroup_widget_areas[repeat]['dataset_dropdown'].value == "RNA": 
            if change not in subgroup_widget_areas[repeat]['variables_combobox'].options:
                subgroup_output_areas[repeat]['subgroup_maker1_info'].clear_output()
                subgroup_output_areas[repeat]['subgroup_maker2_info'].clear_output()
                return

        ##### Step 01 
        # Apply the 0 and 1 labels to the event observed column and filter those values 
        if event_observation_tagsinput0.value and event_observation_tagsinput1.value:
            for tag in event_observation_tagsinput0.value:
                KM_data_1var[event_observation_dropdown.value] = KM_data_1var[event_observation_dropdown.value].replace(tag, "0")
            for tag in event_observation_tagsinput1.value:
                KM_data_1var[event_observation_dropdown.value] = KM_data_1var[event_observation_dropdown.value].replace(tag, "1")

            # Filter the selected 0/1 event labels, transform column to integers and keep only the common first 3 columns 
            KM_data_1var = KM_data_1var.loc[KM_data_1var[event_observation_dropdown.value].isin(["0", "1"])]
            KM_data_1var[event_observation_dropdown.value] = KM_data_1var[event_observation_dropdown.value].astype(int)

        # Log the current status of KM_data_1var
        logger.info(f"[Subgrouping 1st step] The user selected to label -{str(event_observation_tagsinput0.value)}- as 0, and -{str(event_observation_tagsinput1.value)}- as 1. \n")
        logger.info(f"[Subgrouping 1st step] Apply 0/1 labels to column {event_observation_dropdown.value} on KM_data_1var: \n {KM_data_1var.iloc[:15, :10].to_string()} \n")
        logger.info(f"[Subgrouping 1st step] Data types of KM_data_1var columns: \n {KM_data_1var.dtypes.to_string()} \n\n")

        ##### Step 02
        # Look for the selected column in either df, as it is not specified within this function
        if change in df_clinical.columns:
            # Keep only the working columns, log it and extract the column to plot the values
            KM_data_1var = KM_data_1var[['PATIENT_ID', time_to_event_dropdown.value, event_observation_dropdown.value, change]]          
            logger.info(f"[Subgrouping 2nd step] The column {change} -{KM_data_1var.dtypes[change]} dtype- from df_clinical was selected to make subgroups. \n")
            column_data[repeat] = KM_data_1var[change].copy()
            
        # If the column is in df_RNA, joining is required to combine it with the clinical columns
        elif df_RNA is not None and change in df_RNA.columns:
            # Keep only the working columns from both dfs, log it and extract the column to plot the values
            KM_data_1var = KM_data_1var[['PATIENT_ID', time_to_event_dropdown.value, event_observation_dropdown.value]]
            df_RNA2 = df_RNA[['PATIENT_ID', change]]
            KM_data_1var = KM_data_1var.merge(df_RNA2, on='PATIENT_ID', how='inner')
            logger.info(f"[Subgrouping 2nd step] The column {change} -{KM_data_1var.dtypes[change]} dtype- from df_RNA was selected to make subgroups. \n")
            column_data[repeat] = KM_data_1var[change].copy()

        # Log the current status of KM_data_1var 
        logger.info(f"[Subgrouping 2nd step] Keep relevant columns of KM_data_1var and only rows with 0/1 event labels: \n {KM_data_1var.iloc[:15, :10].to_string()} \n")
        logger.info(f"[Subgrouping 2nd step] Data types of KM_data_1var columns: \n {KM_data_1var.dtypes.to_string()} \n\n")

        ##### Step 03
        # Make and display a bar chart for text columns showing the counts for unique values
        if column_data[repeat].dtype == 'object':
            value_counts = column_data[repeat].value_counts(dropna=False)
            fig, ax = plt.subplots(figsize=(5, 3))
            value_counts.plot(kind='bar', color=['indigo', 'khaki', 'lightblue', 'salmon', 'sienna', 'silver', 'aquamarine', 'coral', 'teal', 'olive'])
            ax.set_xlabel(change)
            ax.set_ylabel('Count')
            ax.set_title(f'Unique Value Counts for {change}')
            plt.xticks(rotation=45)

            # Clear the specific space to show the new plot
            subgroup_output_areas[repeat]['subgroup_maker2_info'].clear_output()
            with subgroup_output_areas[repeat]['subgroup_maker1_info']:
                subgroup_output_areas[repeat]['subgroup_maker1_info'].clear_output()
                plt.show()
                display(HTML('<span style="color: red;">This plot shows rows with the 0 and 1 events especified above!</span>'))
        
        # Make and display a histogram of frequencies for numerical columns
        else:
            fig, ax = plt.subplots(figsize=(5, 3))
            ax.hist(column_data[repeat], bins='auto', color="darkblue", ec="white")
            ax.set_xlabel(change)
            ax.set_ylabel('Frequency')
            ax.set_title(f'Histogram for {change}')

            # Clear the specific space to show the new plot
            subgroup_output_areas[repeat]['subgroup_maker2_info'].clear_output()
            with subgroup_output_areas[repeat]['subgroup_maker1_info']:
                subgroup_output_areas[repeat]['subgroup_maker1_info'].clear_output()
                plt.show()
                display(HTML('<span style="color: red;">This plot shows rows with the 0 and 1 events especified above!</span>'))

        ##### Step 04
        # Pass the information of the current variable KM_data_1var so we can have all variables of interest in KM_data_all
        if repeat == 0:
            # The first variable gets just passed as it is after the steps above
            KM_data_all = KM_data_1var.copy()
        else:
            # Check if the column is already in KM_data_all
            if repeat + 3 in KM_data_all.columns:
                # Replace the values in the existing column
                KM_data_all.iloc[:, repeat + 3] = KM_data_1var[change].copy()
            else:
                # Create a new column and assign the values
                KM_data_all[change] = KM_data_1var[change].copy()
        
            # Rename the column to the desired name
            KM_data_all.rename(columns={repeat + 3: change}, inplace=True)

        # Log the KM_data_all as we add/replace variables/columns of interest 
        logger.info(f"[Subgrouping 2nd step] Updated KM_data_all with columns of interest: \n {KM_data_all.iloc[:15, :10].to_string()} \n")
        
        # After the plot is made trigger the slider function to show two (the minimum) widgets in the box 
        subgroup_widget_areas[repeat]['subgroup_number_slider'].value = 2

###################################################################################################

# Function to display the output of group_number slider (widget boxes)
def group_number_selection_handler(change, repeat):
    st.write("Feature 3 pending")
    return
    # This function uses the global variable column_data[repeat] created in the function below
    global subgroup_boxes, subgroup_tagsinput, subgroup_floatrangeslider

    # When the default value on the variable dropdown or combobox is selected, do not display subgrouping options
    if subgroup_widget_areas[repeat]['variables_dropdown'].value == 'Click here to select...':
        if df_RNA is None or subgroup_widget_areas[repeat]['variables_combobox'].value == '':
            with subgroup_output_areas[repeat]['subgroup_maker2_info']:
                subgroup_output_areas[repeat]['subgroup_maker2_info'].clear_output()
                display(HTML('<span style="color: red;">Choose a variable first!</span>'))
            return
    
    # Use tags to specify the desired groups for text columns (min 2, max 10 groups)
    if change>1 and column_data[repeat].dtype == 'object':
        logger.info(f"[Subgrouping 3rd step] The user selected to make {change} subgroups with tags input labels. \n")

        # Make as many tags input widgets and labels as selected and put them in their corresponding box
        unique_values = np.ndarray.tolist(column_data[repeat].unique())
        subgroup_tagsinput = [widgets.TagsInput(allowed_tags=unique_values, description=f'Group {i+1}') for i in range(change)]
        subgroup_tagsinput_labels = [widgets.Label(value=f'Group {i+1}') for i in range(change)]
        subgroup_boxes[repeat] = VBox([HBox([label, tagsinput]) for label, tagsinput in zip(subgroup_tagsinput_labels, subgroup_tagsinput)])

        # Clear the specific space to show the widget box
        with subgroup_output_areas[repeat]['subgroup_maker2_info']:
            subgroup_output_areas[repeat]['subgroup_maker2_info'].clear_output()
            display(subgroup_boxes[repeat])
            
    # Use range sliders to specify the desired groups for numerical columns (min 2, max 10 groups)
    elif change>1:
        logger.info(f"[Subgrouping 3rd step] The user selected to make {change} subgroups with float range sliders. \n")
        
        # Make as many float range slider widgets as selected and put them in their corresponding box
        cleaned_column_data = column_data[repeat].dropna()
        min_value = cleaned_column_data.min()
        max_value = cleaned_column_data.max()
        subgroup_floatrangeslider = [widgets.FloatRangeSlider(min=min_value, max=max_value, step=0.01, description=f'Group {i + 1}') for i in range(change)]
        subgroup_boxes[repeat] = VBox(subgroup_floatrangeslider)

        # Clear the specific space to show the widget box
        with subgroup_output_areas[repeat]['subgroup_maker2_info']:
            subgroup_output_areas[repeat]['subgroup_maker2_info'].clear_output()
            display(subgroup_boxes[repeat]) 
    
    # Tell the user to remove the current variable if they want one group
    else:
        with subgroup_output_areas[repeat]['subgroup_maker2_info']:
            subgroup_output_areas[repeat]['subgroup_maker2_info'].clear_output()
            display(HTML('<span style="color: red;">For 1 group just remove this variable!</span>'))

###################################################################################################

# Function to feed the appropriate selections to the KM_analysis function 
def pass_KM_parameters():
    
    # Get the required variables from the session state
    df_clinical = st.session_state.get("df_clinical")
    logger = st.session_state.get("logger")
    time_to_event_selection = st.session_state.get("time_to_event_selection")
    event_observation_selection = st.session_state.get("event_observation_selection")
    event_0 = st.session_state.get("event_0")
    event_1 = st.session_state.get("event_1")
    subgroup_buttons_selection = st.session_state.get("subgroup_buttons_selection")
    CI_checkbox = st.session_state.get("CI_checkbox")
    move_labels_checkbox = st.session_state.get("move_labels_checkbox")
    at_risk_checkbox = st.session_state.get("at_risk_checkbox")
    sample_fraction = st.session_state.get("sample_fraction")
    KM_plot_area = st.session_state["widget_and_output_areas"][9]

    # Handle missing and required widget inputs before trying to do anything
    if time_to_event_selection == "Click here to select...":
        with KM_plot_area:
            KM_plot_area.empty()
            st.warning("First select the time to event column!!")
            st.stop()
    elif event_observation_selection == "Click here to select...":
        with KM_plot_area:
            KM_plot_area.empty()
            st.warning("First select the event observation column!!")
            st.stop()
    elif event_0 == [] or event_1 == []:
        with KM_plot_area:
            KM_plot_area.empty()
            st.warning("First select the values to label as 0 and 1 (No event, event)!!")
            st.stop()

    # If no subgrouping is required, apply the event tags and pass the data to KM_analysis
    if subgroup_buttons_selection == 'None':
        
        # Apply the selected labels on the event observation column 
        KM_data = df_clinical.copy()
        for tag in event_0:
            KM_data[event_observation_selection] = KM_data[event_observation_selection].replace(tag, "0")
        for tag in event_1:
            KM_data[event_observation_selection] = KM_data[event_observation_selection].replace(tag, "1")            

        # Log the current status of KM_data
        logger.info(f"[No subgroups 1st step] The user selected to label -{str(event_0)}- as 0, and -{str(event_1)}- as 1. \n")
        logger.info(f"[No subgroups 1st step] Apply 0/1 labels to column {event_observation_selection} on KM_data: \n {KM_data.iloc[:15, :10].to_string()} \n")
        logger.info(f"[No subgroups 1st step] Data types of KM_data columns: \n {KM_data.dtypes.to_string()} \n\n")
                
        # Filter out non-desired values and convert column to numbers for the KM Fitter
        KM_data = KM_data[['PATIENT_ID', time_to_event_selection, event_observation_selection]]
        KM_data = KM_data.loc[KM_data[event_observation_selection].isin(["0", "1"])]
        KM_data = KM_data.dropna(subset=[time_to_event_selection])
        KM_data[event_observation_selection] = KM_data[event_observation_selection].astype(int)

        # Log the current status of KM_data
        logger.info(f"[No subgroups 2nd step] Keep relevant columns of KM_data and only rows with 0/1 event labels: \n {KM_data.head(15).to_string()} \n")
        logger.info(f"[No subgroups 2nd step] Data types of KM_data columns: \n {KM_data.dtypes.to_string()} \n\n")
        
        # Pass the input parameters to the KM_analysis function and get back the KM object
        KM_subgroups = []           
        KM_analysis_output = KM_analysis(KM_data, KM_subgroups)
        
        # Make a plot with altair for the KM estimate obtained
        plt.figure(figsize=(10, 6))
        KM_analysis_output.plot(ci_show=CI_checkbox, legend=False, at_risk_counts=at_risk_checkbox, 
                                iloc=slice(0, int(len(KM_analysis_output.survival_function_) * sample_fraction)))
        if not at_risk_checkbox:
            plt.xlabel("Time (Months)")
        plt.ylabel("Survival Probability")
        plt.title("Kaplan-Meier Estimate")
        
    # If subgroups were selected, apply the corresponding tags or ranges
    else:
        ######CHECK where KM_data_all and variable_repeats come from

        # Log the current status of KM_data_working
        logger.info(f"[Subgrouping 3rd step] Dataset KM_data_all before applying subgrouping labels: \n {KM_data_all.head(15).to_string()} \n")
        logger.info(f"[Subgrouping 3rd step] Data types of KM_data_all before applying subgrouping labels: \n {KM_data_all.dtypes.to_string()} \n\n")
        
        # If the subgrouping changes are made multiple times, apply them to a copy of the original df
        KM_data_working = KM_data_all.copy()

        # Create an empty dictionary to store the mapping for each variable to reassign real group names
        correct_group_labels = [{} for i in range(5)]

        # Only iterate through the last number of variables in the slider (in case the user had more before)
        for repeat in range(variable_repeats):

            # To apply the tags in tagsinput we check if the VBoxes have repeats of label+tagsinput or floatrangesliders alone
            if isinstance(subgroup_boxes[repeat].children[0], widgets.FloatRangeSlider):
                
                # Create a new column to store the group labels (original is numbers, new one will be text and next to the original)
                KM_data_working.insert(repeat+4, 'TextSubgroup', '')

                # Iterate through the float range sliders
                for i, floatslider in enumerate(subgroup_boxes[repeat].children):
                    # Retrieve the range selection and corresponding label
                    subgroup_range = floatslider.value
                    subgroup_label = floatslider.description
                   
                    # Get the indices of rows within the range selected
                    subgroup_rows = (KM_data_working.iloc[:, repeat+3] >= subgroup_range[0]) & (KM_data_working.iloc[:, repeat+3] < subgroup_range[1])
                
                    # Assign the subgroup label to the matching rows
                    KM_data_working.loc[subgroup_rows, 'TextSubgroup'] = subgroup_label

                    # Add the correct label to the dictionary
                    correct_group_labels[repeat][subgroup_label] = f"{subgroup_range[0]:.2f} to {subgroup_range[1]:.2f}"
                
                # Remove rows where the subgroup label is not assigned
                KM_data_working = KM_data_working[KM_data_working['TextSubgroup'] != '']
                variable_column_name = KM_data_working.columns[repeat+3]
                KM_data_working.drop(variable_column_name, axis=1, inplace=True)
                KM_data_working.rename(columns={'TextSubgroup': variable_column_name}, inplace=True)

                # Log the ranges corresponding to each subgroup
                log_string = " - ".join([f"Group {i+1}: {slider.value[0]:.2f} to {slider.value[1]:.2f}" for i, slider in enumerate(subgroup_boxes[repeat].children)])
                logger.info(f"[Subgrouping 3rd step] Subgrouping labels applied to variable {repeat+1}---> {log_string} \n")

            # The HBoxes containing tagsinput widgets and labels do not have the attribute lenght
            else:
                
                # Retrieve the labels and elements to label
                subgroup_selections = [tagsHBox.children[1].value for tagsHBox in subgroup_boxes[repeat].children]
                subgroup_labels = [tagsHBox.children[0].value for tagsHBox in subgroup_boxes[repeat].children]
                label_content_pairs = []
                
                # Iterate through the subgroup_selections list
                for i, tags_list in enumerate(subgroup_selections):
                    subgroup_elements = tags_list
    
                    # Generate a mapping of unique values to group labels
                    element_to_label = {element: subgroup_labels[i] for element in subgroup_elements}

                    # Add the group label to its list and the correct label to the dictionary
                    label_content_pairs.extend([f"{element}: {subgroup_labels[i]}" for element in subgroup_elements])
                    correct_group_labels[repeat][subgroup_labels[i]] = '+'.join(subgroup_elements)
                    
                    # Replace the subgroup elements with the new labels
                    variable_column_name = KM_data_working.columns[repeat+3]
                    KM_data_working[variable_column_name] = KM_data_working[variable_column_name].replace(element_to_label)
            
                # Filter out rows with new subgroup labels and log the labels selected
                KM_data_working = KM_data_working[KM_data_working[variable_column_name].isin(subgroup_labels)]
                log_string = "  -  ".join(label_content_pairs)
                logger.info(f"[Subgrouping 3rd step] Subgrouping labels applied to variable {repeat+1}---> {log_string} \n")
            
            # Log the updated df
            logger.info(f"[Subgrouping 3rd step] Dataset KM_data_working after applying subgrouping labels: \n {KM_data_working.head(15).to_string()} \n")
        ########
        # Once all labels have been applied to each column, make the subgroups

        # Get the column indices for the extra columns
        extra_column_indices = range(3, 3 + variable_repeats)
        
        # Get the unique values for each extra column
        extra_column_unique_values = [KM_data_working.iloc[:, i].unique() for i in extra_column_indices]
        
        # Create an empty dictionary to store the subsets
        KM_subgroups = {}
        
        # Generate all possible combinations of unique values from the extra columns
        combinations = [[]]
        for values in extra_column_unique_values:
            combinations = [sublist + [value] for sublist in combinations for value in values]
        
        # Iterate through each combination of unique values
        for combination in combinations:
            # Create a subset of KM_data_working for the current combination
            subset = KM_data_working.copy()
            
            # Filter the rows based on the current combination
            for i, index in enumerate(extra_column_indices):
                subset = subset[subset.iloc[:, index] == combination[i]]
            
            # Add the subset to the KM_subgroups dictionary
            KM_subgroups[tuple(combination)] = subset

        # Log the subgroups created
        logger.info(f"[Subgrouping 3rd step] Subgroups made from the dataset:\n")
        for combination, subgroup in KM_subgroups.items():
            logger.info(f"Subgroup label: {combination}")
            logger.info(f"\n{subgroup.head(10)}\n")
        
        ########
                
        # Finally, pass the input parameters to the KM_analysis function and get back the KM object
        KM_analysis_output = KM_analysis(KM_data_working, KM_subgroups)

        # Reassign the real/correct labels (they are as Group X, and we will correct to the actual tag or range)
        # Iterate through each key-value pair in the KM_analysis_output dictionary
        for old_key, KM_object in list(KM_analysis_output.items()):
            # Create a list to store the new key for this combination
            new_key = []
        
            # Iterate through each element of the old key (a tuple of strings)
            for i, label in enumerate(old_key):
                # Retrieve the correct label from the corresponding dictionary in correct_group_labels
                new_key.append(correct_group_labels[i].get(label, label))
        
            # Convert the new key (list) to a single string
            new_key = ', '.join(new_key)
        
            # Replace the current key with the corrected_key in the KM_analysis_output dictionary
            KM_analysis_output[new_key] = KM_analysis_output.pop(old_key)

        # Plot the estimates of all KMF objects 
        with KM_plot_area:
            KM_plot_area.empty()     
            plt.figure(figsize=(10, 6))
            for label, KM_object in KM_analysis_output.items():
                KM_object.plot(label=label, ci_show=CI_checkbox.value, iloc=slice(0, int(len(KM_object.survival_function_) * sample_fraction)))
            plt.xlabel('Time (Months)')
            plt.ylabel('Probability')
            plt.title('Kaplan-Meier Estimates')
            plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left') if move_labels_checkbox.value else plt.legend()
            if at_risk_checkbox.value:
                add_at_risk_counts(*KM_analysis_output.values(), labels=list(KM_analysis_output.keys()), ax=plt.gca())
            plt.show()

    # Save the KMF objects made and the figure diplayed to the session state
    st.session_state["KM_analysis_output"] = KM_analysis_output
    figure_bytes = io.BytesIO()
    plt.savefig(figure_bytes, format='jpg', dpi=600, bbox_inches='tight')
    figure_bytes.seek(0)
    st.session_state['logged_figure'] = figure_bytes

    return plt.gcf()

###################################################################################################

def KM_analysis(KM_data, KM_subgroups):
    
    # Unpack the input parameters provided
    current_time_column = KM_data.columns[1]
    current_event_column = KM_data.columns[2]

    # Get the required variables from the session state
    logger = st.session_state.get("logger")
    subgroup_buttons_selection = st.session_state.get("subgroup_buttons_selection")

    # Use the whole dataset when no groups were made
    if subgroup_buttons_selection == 'None':

        # Create a single KaplanMeierFitter object
        KMF_object = KaplanMeierFitter()

        # Generate the plot using the specified columns
        KMF_object.fit(durations=KM_data[current_time_column], event_observed=KM_data[current_event_column])

        # Log part of the curve to verify the data was passed correctly
        logger.info(f"[No Subgroups 3rd step] The KM Fitter succesfully calculated the probabilities and made the plot. \n")
        logger.info(f"[No Subgroups 3rd step] Calculated survival function: \n {KMF_object.survival_function_.head(7).to_string()} \n ... \n {KMF_object.survival_function_.tail(7).to_string()} \n\n")

    # Make a fit for every subset provided (based on the number of groups and subgroups made
    else:
        
        # Sort the subgroups in alphabetical order to plot them in the same order and colour
        KM_subgroups = OrderedDict(sorted(KM_subgroups.items()))
        
        # Create an empty dictionary to store the KaplanMeierFitter objects
        KMF_object = {}
        logger.info(f"[Subgrouping 4th step] The KM Fitter succesfully calculated the probabilities. \n")
        
        # Create KaplanMeierFitter objects for each subgroup in KM_subgroups
        for label, subset in KM_subgroups.items():
            kmf = KaplanMeierFitter()
            kmf.fit(durations=subset[current_time_column], event_observed=subset[current_event_column])
            KMF_object[label] = kmf

            # Log part of the curve to verify the data was passed correctly
            logger.info(f"[Subgrouping 4th step] Calculated survival function of: {label}")
            logger.info(f"\n {kmf.survival_function_.head(7).to_string()} \n ... \n {kmf.survival_function_.tail(7).to_string()} \n\n")
        
    return KMF_object

###################################################################################################

def save_KM_results(generate_plot_button):

    # Get the required variables from the session state
    KM_analysis_output = st.session_state.get("KM_analysis_output")
    logger = st.session_state.get("logger")
    file_count = st.session_state.get("file_count", 1)
    col_2_row_15 = st.session_state["widget_and_output_areas"][12]
    col_3_row_15 = st.session_state["widget_and_output_areas"][13]

    # File names for Excel and plot
    file_count_str = str(file_count).zfill(2)  # Convert to a 2-digit zero-padded string
    excel_filename = f"KM_results_{file_count_str}.xlsx"
    plot_filename = f"KM_results_{file_count_str}.jpg"

    ###################### Make the plot available for download

    # Get the data from the plot already being displayed
    figure_bytes = st.session_state.get('logged_figure')

    ###################### Make the Excel file available for download

    # Make a new excel if none has been made or the user explicitely wants a new one
    if "logged_excel" not in st.session_state or generate_plot_button:    
        
        # Create a new Excel workbook and remove the default Sheet
        workbook = openpyxl.Workbook()
        workbook.remove(workbook['Sheet'])
        
        # Prepare the data to be processed (single KM object or list of KM objects)
        if isinstance(KM_analysis_output, dict):
            KM_objects_to_process = [{"label": f"KM_Subgroup_{i+1}", "KM_object": KM_object} for i, (label, KM_object) in enumerate(KM_analysis_output.items())]
            real_labels = [f"KM_Subgroup_{i+1}: {label}" for i, (label, KM_object) in enumerate(KM_analysis_output.items())]
        else:
            # If KM_analysis_output is a single KM object, add it to the list as a dictionary with a general label
            KM_objects_to_process = [{"label": "KM_Dataset", "KM_object": KM_analysis_output}]
            real_labels = ["KM_Dataset: Whole dataset - No subgroups"]

        # Process all KM curves/objects the same way 
        for index, data in enumerate(KM_objects_to_process):
            
            # Create a sheet per KM object
            sheet = workbook.create_sheet(title=data["label"])

            # Write what the curve/object corresponds to
            sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=13)
            sheet.cell(row=2, column=1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            sheet.cell(row=2, column=1, value=real_labels[index]).font = openpyxl.styles.Font(bold=True, size=16)

            # Get the tables from the KMF object
            event_table = data["KM_object"].event_table
            survival_function = pd.DataFrame({"Time": data["KM_object"].survival_function_.index,
                                            "Survival Probability": np.ravel(data["KM_object"].survival_function_.values)})
            confidence_interval = data["KM_object"].confidence_interval_
            median_survival_time = data["KM_object"].median_survival_time_

            # Write the tables to the Excel sheet
            tables = [event_table, survival_function, confidence_interval, median_survival_time]
            table_names = ["Event Table", "Survival Function", "Confidence Intervals", "Median Survival Time"]
            table_column_numbers = [1, 7, 10, 13]
            
            # Write all tables to the sheet
            for col_index, (table, table_name) in enumerate(zip(tables, table_names)):
                # Define the current column number for the table
                current_column = table_column_numbers[col_index]
        
                # Set the header for the current table
                sheet.cell(row=4, column=current_column, value=table_name).font = openpyxl.styles.Font(bold=True)
        
                if isinstance(table, pd.DataFrame):
                    # If the table is a DataFrame, convert it to a NumPy array
                    rows = table.to_numpy()
        
                    for row_index, row in enumerate(rows):
                        # Write the data from the DataFrame to the Excel sheet
                        for col_offset, value in enumerate(row):
                            sheet.cell(row=row_index + 6, column=current_column + col_offset, value=value)  
                else:
                    # If the table is not a DataFrame, write the single value to the Excel sheet
                    sheet.cell(row=5, column=current_column, value=table)

            ##### Extra worksheet formatting 
            # Merge and center table titles
            sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=5)
            sheet.cell(row=4, column=1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            sheet.merge_cells(start_row=4, start_column=7, end_row=4, end_column=8)
            sheet.cell(row=4, column=7).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            sheet.merge_cells(start_row=4, start_column=10, end_row=4, end_column=11)
            sheet.cell(row=4, column=10).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # Write column titles and center them
            sheet.cell(row=5, column=1, value="Removed")
            sheet.cell(row=5, column=2, value="Observed")
            sheet.cell(row=5, column=3, value="Censored")
            sheet.cell(row=5, column=4, value="Entrance")
            sheet.cell(row=5, column=5, value="At Risk")
            sheet.cell(row=5, column=7, value="Time")
            sheet.cell(row=5, column=8, value="Probability")
            sheet.cell(row=5, column=10, value="Lower Bound")
            sheet.cell(row=5, column=11, value="Upper Bound")
            for cell in ["A5", "B5", "C5", "D5", "E5", "G5", "H5", "J5", "K5"]:
                sheet[cell].alignment = openpyxl.styles.Alignment(horizontal='center')
            
            # Adjust some column widths
            for column, width in zip(['H', 'J', 'K', 'M'], [10, 12, 12, 22]):
                sheet.column_dimensions[column].width = width
            #####
        
        # Save the Excel file to the session state
        excel_bytes = io.BytesIO()
        workbook.save(excel_bytes)
        excel_bytes.seek(0)
        st.session_state['logged_excel'] = excel_bytes
    else:
        # If this rerun does not require making a new excel, use the logged file
        excel_bytes = st.session_state.get("logged_excel")

    ###################### Show the download buttons for the current data
    
    with col_2_row_15:
        download_plot = st.download_button(label='Download Plot', data=figure_bytes, 
                           file_name=plot_filename, type='primary', mime='image/png')
    with col_3_row_15:
        download_excel = st.download_button(label='Download Raw Data', data=excel_bytes, 
                           file_name=excel_filename, type='primary',
                           mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # Log the download of any file
    if download_plot:
        logger.info(f"A KM plot with the name {plot_filename} has been downloaded \n")
    if download_excel:
        logger.info(f"An excel file with the name {excel_filename} has been downloaded \n")
    # When something is downloaded, increment the file count to avoid multiple files with the same name
    if download_plot or download_excel:
        st.session_state["file_count"] = file_count + 1

###################################################################################################
######################################### Flow control ############################################

# Clear the saved df's in case the user wants to upload other files
if restart_button:
    st.session_state.clear()

# Start the processing when the user clicks on the button
if start_button or "flow_control_1" in st.session_state:
    
    # However, proceed only when there have been files uploaded
    if uploaded_files:

        # Create a control for reruns to display the widgets once the user has uploaded
        # files and clicked the Begin button at least once
        st.session_state["flow_control_1"] = True

        # Load the file(s)
        if "df_clinical" not in st.session_state:
            load_input_files(uploaded_files)

        # Process the file(s) to get the information for the main widgets
        if "time_to_event" not in st.session_state and "event_observation" not in st.session_state:
            file_preprocessing()
        
        # Prepare and display the widgets to generate the plots
        widget_preparation()

    else:
        # Display an warning message because we are missing the clinical file
        st.sidebar.warning("Please upload your clinical data file!")

###################################################################################################
